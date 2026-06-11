"""Computer Resource Monitor - standalone once-off run script.

This is the notebook monitor.ipynb as a plain Python script, for running the whole job in
one go. From PowerShell:

    python monitor.py          # real run, 5 minute interval for 24 hours
    python monitor.py --demo   # quick demonstration, about 30 seconds

It checks dependencies, collects resources, raises the concluding alarm that benchmarks
against normal, and builds the trend charts, writing everything to the outputs folder.

Generated from monitor.ipynb. Edit the notebook and regenerate rather than editing this
file by hand.
"""

from datetime import timedelta
from pathlib import Path

# ----- Sampling -----

# How long to wait between samples.
SAMPLE_INTERVAL = timedelta(minutes=5)

# How long the whole run lasts. After this elapses the loop stops.
TOTAL_DURATION = timedelta(hours=24)

# ----- Output file paths (do not change these names) -----

DASHBOARD_PATH = Path("outputs") / "dashboard" / "dashboard.xlsx"
LOG_PATH = Path("outputs") / "log" / "log.xlsx"
CHARTS_DIR = Path("outputs") / "charts"
BENCHMARK_PATH = Path("outputs") / "benchmark.json"

# Make sure the output folders exist.
DASHBOARD_PATH.parent.mkdir(parents=True, exist_ok=True)
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
CHARTS_DIR.mkdir(parents=True, exist_ok=True)

# ----- Alarm thresholds (a sample is flagged when a value crosses these) -----

CPU_ALARM_PERCENT = 90          # overall CPU usage above this is abnormal
MEMORY_ALARM_PERCENT = 90       # memory usage above this is abnormal
DISK_ALARM_PERCENT = 90         # any drive above this is abnormal
GPU_ALARM_PERCENT = 95          # GPU usage above this is abnormal
TEMPERATURE_ALARM_C = 85        # CPU temperature above this is abnormal
BATTERY_LOW_PERCENT = 15        # battery below this, on battery power, is abnormal

# Windows services that should always be running. A stopped one is flagged.
REQUIRED_SERVICES = ["Dnscache", "Schedule", "EventLog", "Winmgmt"]

# ----- Statistical baseline settings -----

# Fraction of the run treated as normal conditions to learn the benchmark from.
BASELINE_FRACTION = 0.25
# How many standard deviations away from the normal mean counts as a deviation.
BASELINE_STD_MULTIPLIER = 3

# Derived: total number of samples this run will take.
SAMPLE_COUNT = int(TOTAL_DURATION / SAMPLE_INTERVAL)

print("Sample interval :", SAMPLE_INTERVAL)
print("Total duration  :", TOTAL_DURATION)
print("Planned samples :", SAMPLE_COUNT)
print("Dashboard file  :", DASHBOARD_PATH)
print("Log file        :", LOG_PATH)
print("Charts folder   :", CHARTS_DIR)
print("Benchmark file  :", BENCHMARK_PATH)
print("Alarm limits    : CPU", CPU_ALARM_PERCENT, "Memory", MEMORY_ALARM_PERCENT,
      "Disk", DISK_ALARM_PERCENT, "GPU", GPU_ALARM_PERCENT,
      "Temp", TEMPERATURE_ALARM_C, "Battery low", BATTERY_LOW_PERCENT)
print("Required services:", REQUIRED_SERVICES)


import importlib
import importlib.metadata as metadata
import subprocess
import sys

# Import name mapped to its pip name and the minimum version the notebook needs.
# matplotlib is used only at the end of the run to draw the trend graphs.
REQUIRED_PACKAGES = {
    "psutil": {"pip": "psutil", "min": "5.9.6"},
    "openpyxl": {"pip": "openpyxl", "min": "3.0.0"},
    "matplotlib": {"pip": "matplotlib", "min": "3.0.0"},
}


def _version_tuple(text):
    """Turn a version string like '5.9.6' into a comparable tuple of integers."""
    parts = []
    for chunk in text.split("."):
        digits = ""
        for char in chunk:
            if char.isdigit():
                digits += char
            else:
                break
        parts.append(int(digits) if digits else 0)
    return tuple(parts)


def _pip_install(spec):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", spec])


def ensure_packages(packages):
    """Install or upgrade each package so it is present and meets the minimum version."""
    print("Python version:", sys.version.split()[0])
    print("-" * 50)

    changed = False
    for import_name, info in packages.items():
        pip_name = info["pip"]
        minimum = info["min"]
        try:
            current = metadata.version(pip_name)
        except metadata.PackageNotFoundError:
            current = None

        if current is None:
            print("NOT INSTALLED  :", pip_name, "(need >=", minimum + ")")
            _pip_install(pip_name + ">=" + minimum)
            changed = True
        elif _version_tuple(current) < _version_tuple(minimum):
            print("OUTDATED       :", pip_name, current, "(need >=", minimum + ")")
            _pip_install(pip_name + ">=" + minimum)
            changed = True
        else:
            print("OK             :", pip_name, current)

    print("-" * 50)
    if changed:
        print("Packages were installed or upgraded.")
        print("Restart the kernel and run the notebook again from the top.")
    else:
        print("All required packages are present and up to date.")


ensure_packages(REQUIRED_PACKAGES)


import psutil
from datetime import datetime


def collect_cpu(timestamp):
    """Return overall and per-core CPU usage for one sample, in plain labels."""
    per_core = psutil.cpu_percent(interval=1, percpu=True)
    overall = round(sum(per_core) / len(per_core), 1) if per_core else 0.0

    row = {
        "Time": timestamp,
        "Overall CPU Usage (%)": overall,
        "Number of CPU Cores": len(per_core),
    }
    for index, value in enumerate(per_core):
        # Cores numbered from 1 so the labels read naturally.
        row["Core {} Usage (%)".format(index + 1)] = value
    return row


import subprocess
from datetime import datetime


def _run_powershell(command, timeout=15):
    """Run a PowerShell command and return its text output, or None on failure."""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", command],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass
    return None


def collect_temperature(timestamp):
    """Return the CPU temperature in Celsius, or Unavailable if the sensor is blocked."""
    command = (
        "Get-CimInstance -Namespace root/wmi "
        "-ClassName MSAcpi_ThermalZoneTemperature -ErrorAction Stop "
        "| Select-Object -ExpandProperty CurrentTemperature"
    )
    output = _run_powershell(command)

    temperature = "Unavailable"
    if output:
        readings = []
        for line in output.splitlines():
            line = line.strip()
            if line.isdigit():
                # WMI reports tenths of a Kelvin. Convert to Celsius.
                readings.append(int(line) / 10.0 - 273.15)
        if readings:
            temperature = round(sum(readings) / len(readings), 1)

    return {
        "Time": timestamp,
        "CPU Temperature (°C)": temperature,
    }


import psutil
from datetime import datetime

BYTES_PER_GB = 1024 ** 3


def to_gb(value):
    """Convert a byte count to gigabytes, rounded to two decimals."""
    return round(value / BYTES_PER_GB, 2)


def collect_ram(timestamp):
    """Return total, used, available memory in GB and the used percentage."""
    memory = psutil.virtual_memory()
    return {
        "Time": timestamp,
        "Total Memory (GB)": to_gb(memory.total),
        "Used Memory (GB)": to_gb(memory.used),
        "Available Memory (GB)": to_gb(memory.available),
        "Memory Usage (%)": memory.percent,
    }


import psutil
from datetime import datetime

BYTES_PER_GB = 1024 ** 3


def _gb(value):
    return round(value / BYTES_PER_GB, 2)


def collect_disk(timestamp):
    """Return one row per drive with total, used, free space and used percentage."""
    rows = []
    for partition in psutil.disk_partitions(all=False):
        try:
            usage = psutil.disk_usage(partition.mountpoint)
        except Exception:
            # Empty CD drive, card reader, or a drive psutil cannot read on Windows
            # (this can raise PermissionError, OSError, or SystemError). Skip it so
            # one bad drive does not stop the sample.
            continue
        rows.append({
            "Time": timestamp,
            "Drive": partition.device,
            "Total Space (GB)": _gb(usage.total),
            "Used Space (GB)": _gb(usage.used),
            "Free Space (GB)": _gb(usage.free),
            "Disk Usage (%)": usage.percent,
        })
    return rows


import psutil
from datetime import datetime


def _format_time_left(seconds):
    """Turn a seconds-remaining value into a plain reading."""
    if seconds == psutil.POWER_TIME_UNLIMITED:
        return "Plugged in"
    if seconds == psutil.POWER_TIME_UNKNOWN or seconds is None or seconds < 0:
        return "Unknown"
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return "{} hours {} minutes".format(hours, minutes)


def collect_battery(timestamp):
    """Return battery charge, plugged-in state and estimated time remaining."""
    battery = psutil.sensors_battery()
    if battery is None:
        return {
            "Time": timestamp,
            "Battery Charge (%)": "No battery",
            "Power Plugged In": "No battery",
            "Estimated Time Remaining": "No battery",
        }
    return {
        "Time": timestamp,
        "Battery Charge (%)": round(battery.percent, 1),
        "Power Plugged In": "Yes" if battery.power_plugged else "No",
        "Estimated Time Remaining": (
            "Plugged in" if battery.power_plugged
            else _format_time_left(battery.secsleft)
        ),
    }


import psutil
from datetime import datetime

BYTES_PER_MB = 1024 ** 2

# Remembers the previous sample's byte counters per interface, so we can report
# the difference rather than the running total. Persists between samples.
_previous_net_counters = {}


def _mb(value):
    return round(value / BYTES_PER_MB, 3)


def collect_network(timestamp):
    """Return per-interface data sent and received since the last sample, plus totals."""
    rows = []
    counters = psutil.net_io_counters(pernic=True)
    for interface, stats in counters.items():
        previous = _previous_net_counters.get(interface)
        if previous is None:
            sent_delta = 0
            recv_delta = 0
        else:
            sent_delta = max(0, stats.bytes_sent - previous.bytes_sent)
            recv_delta = max(0, stats.bytes_recv - previous.bytes_recv)
        rows.append({
            "Time": timestamp,
            "Interface": interface,
            "Data Sent Since Last Sample (MB)": _mb(sent_delta),
            "Data Received Since Last Sample (MB)": _mb(recv_delta),
            "Total Data Sent (MB)": _mb(stats.bytes_sent),
            "Total Data Received (MB)": _mb(stats.bytes_recv),
        })
        _previous_net_counters[interface] = stats
    return rows


from datetime import datetime

# PowerShell that sums GPU engine utilization and GPU process memory across all
# instances. Each value is printed on its own labelled line, or NA if unreadable.
_GPU_COMMAND = r"""
try {
    $u = ((Get-Counter '\GPU Engine(*)\Utilization Percentage' -ErrorAction Stop).CounterSamples |
          Measure-Object -Property CookedValue -Sum).Sum
} catch { $u = 'NA' }
try {
    $m = ((Get-Counter '\GPU Process Memory(*)\Local Usage' -ErrorAction Stop).CounterSamples |
          Measure-Object -Property CookedValue -Sum).Sum
} catch { $m = 'NA' }
Write-Output ("UTIL=" + $u)
Write-Output ("MEM=" + $m)
"""


def collect_gpu(timestamp):
    """Return GPU usage percent and GPU memory used in MB, or Unavailable."""
    output = _run_powershell(_GPU_COMMAND, timeout=20)

    usage = "Unavailable"
    memory_mb = "Unavailable"
    if output:
        for line in output.splitlines():
            line = line.strip()
            if line.startswith("UTIL="):
                value = line[5:].strip().replace(",", ".")
                try:
                    usage = min(100.0, round(float(value), 1))
                except ValueError:
                    pass
            elif line.startswith("MEM="):
                value = line[4:].strip().replace(",", ".")
                try:
                    memory_mb = round(float(value) / (1024 ** 2), 1)
                except ValueError:
                    pass

    return {
        "Time": timestamp,
        "GPU Usage (%)": usage,
        "GPU Memory Used (MB)": memory_mb,
    }


import psutil
from datetime import datetime


def _format_uptime(seconds):
    """Turn a number of seconds into days, hours and minutes."""
    seconds = int(seconds)
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60
    return "{} days {} hours {} minutes".format(days, hours, minutes)


def collect_uptime(timestamp):
    """Return the boot time and how long the machine has been running."""
    boot = datetime.fromtimestamp(psutil.boot_time())
    uptime_seconds = (timestamp - boot).total_seconds()
    return {
        "Time": timestamp,
        "Last Boot Time": boot.strftime("%Y-%m-%d %H:%M:%S"),
        "Uptime": _format_uptime(uptime_seconds),
    }


import psutil
from datetime import datetime

BYTES_PER_MB = 1024 ** 2

# Number of logical cores. Used to convert per-process CPU into a 0 to 100 share of
# total CPU capacity, the way Task Manager shows it, instead of a per-core figure that
# can climb above 100 on a multi-core machine.
CPU_CORE_COUNT = psutil.cpu_count() or 1


def collect_processes(timestamp):
    """Return one row per running process with cpu, memory and status."""
    rows = []
    for proc in psutil.process_iter(["pid", "name", "cpu_percent", "memory_info", "status"]):
        try:
            info = proc.info
            memory = info.get("memory_info")
            memory_mb = round(memory.rss / BYTES_PER_MB, 1) if memory else 0.0
            name = info.get("name") or "Unknown"
            cpu_share = (info.get("cpu_percent") or 0.0) / CPU_CORE_COUNT
            rows.append({
                "Time": timestamp,
                "Process ID (PID)": info.get("pid"),
                "Process Name": name,
                "CPU Usage (%)": round(cpu_share, 1),
                "Memory Usage (MB)": memory_mb,
                "Status": info.get("status") or "Unknown",
            })
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            # Process ended or is protected by Windows. Skip it.
            continue
    return rows


import psutil
from datetime import datetime


def collect_services(timestamp):
    """Return one row per Windows service with status and start type."""
    rows = []
    for service in psutil.win_service_iter():
        try:
            info = service.as_dict()
            rows.append({
                "Time": timestamp,
                "Service Name": info.get("name") or "Unknown",
                "Display Name": info.get("display_name") or "Unknown",
                "Status": info.get("status") or "Unknown",
                "Start Type": info.get("start_type") or "Unknown",
            })
        except Exception:
            # Service could not be read. Skip it so the sample continues.
            continue
    return rows


import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

# The five sheets used in both the dashboard and the log.
SHEET_NAMES = ["system_metrics", "disk", "network", "processes", "services"]


def _build_datasets(timestamp):
    """Run every collector for one sample and group results by sheet name."""
    system_row = {}
    for collector in (
        collect_cpu,
        collect_temperature,
        collect_ram,
        collect_battery,
        collect_gpu,
        collect_uptime,
    ):
        system_row.update(collector(timestamp))

    return {
        "system_metrics": [system_row],
        "disk": collect_disk(timestamp),
        "network": collect_network(timestamp),
        "processes": collect_processes(timestamp),
        "services": collect_services(timestamp),
    }


def _sheet_is_empty(ws):
    """True when a worksheet has no header row yet.

    This checks the sheet dimensions instead of reading a cell such as ws['A1']. On a
    freshly created sheet, reading a cell instantiates it and pushes the first append
    down to row 2, leaving a blank first row that then breaks later appends. Reading
    max_row and max_column does not touch any cell, so it stays clean.
    """
    return ws.max_row == 1 and ws.max_column == 1


def write_dashboard(datasets):
    """Rebuild the dashboard workbook from scratch and overwrite it (latest snapshot)."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_NAMES:
        ws = wb.create_sheet(title=name)
        rows = datasets.get(name) or []
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(header) for header in headers])
        else:
            ws.append(["Time", "No data this sample"])
    wb.save(DASHBOARD_PATH)


def append_log(datasets):
    """Append this sample to the log workbook, creating it on the first sample."""
    if LOG_PATH.exists():
        wb = load_workbook(LOG_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for name in SHEET_NAMES:
        ws = wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)
        rows = datasets.get(name) or []
        if not rows:
            continue
        if _sheet_is_empty(ws):
            headers = list(rows[0].keys())
            ws.append(headers)
        else:
            headers = [cell.value for cell in ws[1]]
        for row in rows:
            ws.append([row.get(header) for header in headers])
    wb.save(LOG_PATH)


import time
from datetime import datetime, timedelta

from openpyxl import load_workbook


def run_monitor(interval=None, duration=None, fresh_start=True, verbose=True):
    """Sample resources on a fixed interval for a fixed duration, writing both files.

    interval and duration default to the config values (5 minutes, 24 hours). A short
    interval and duration can be passed for testing. fresh_start removes any existing
    dashboard and log first so each run produces one clean log.
    """
    interval = interval if interval is not None else SAMPLE_INTERVAL
    duration = duration if duration is not None else TOTAL_DURATION

    if fresh_start:
        for path in (DASHBOARD_PATH, LOG_PATH):
            if path.exists():
                path.unlink()

    # Reset the network baseline so the first sample's per-sample data starts at zero.
    _previous_net_counters.clear()

    start = datetime.now()
    end = start + duration
    next_time = start
    sample_number = 0

    print("Run started at", start.strftime("%Y-%m-%d %H:%M:%S"))
    print("Interval:", interval, " Duration:", duration)
    print("-" * 50)

    while datetime.now() < end:
        sample_number += 1
        timestamp = datetime.now()
        try:
            datasets = _build_datasets(timestamp)
            write_dashboard(datasets)
            append_log(datasets)
            if verbose:
                system_row = datasets["system_metrics"][0]
                print("Sample {} at {}   CPU {}%   Memory {}%".format(
                    sample_number,
                    timestamp.strftime("%H:%M:%S"),
                    system_row.get("Overall CPU Usage (%)"),
                    system_row.get("Memory Usage (%)"),
                ))
        except Exception as error:
            print("Sample", sample_number, "failed:", error)

        # Keep a steady cadence: sleep until the next scheduled sample time.
        next_time += interval
        sleep_for = (next_time - datetime.now()).total_seconds()
        if sleep_for > 0 and datetime.now() < end:
            time.sleep(sleep_for)

    print("-" * 50)
    print("Run complete.", sample_number, "samples over", datetime.now() - start)


from openpyxl import load_workbook


def read_log_sheet(sheet_name, path=None):
    """Load one sheet of the log into a list of plain dictionaries keyed by header."""
    path = path or LOG_PATH
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, values)) for values in rows[1:]]


def _alarm(time, source, resource, reading, limit, message):
    """Build one plain-language alarm row."""
    return {
        "Time": time,
        "Source": source,
        "Resource": resource,
        "Reading": reading,
        "Limit": limit,
        "Message": message,
    }


def _is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def check_fixed_thresholds():
    """Flag every sample value that crosses a fixed limit from the config section."""
    alarms = []

    for row in read_log_sheet("system_metrics"):
        time = row.get("Time")

        cpu = row.get("Overall CPU Usage (%)")
        if _is_number(cpu) and cpu > CPU_ALARM_PERCENT:
            alarms.append(_alarm(time, "Fixed threshold", "CPU", cpu, CPU_ALARM_PERCENT,
                "CPU usage {}% is above the {}% limit".format(cpu, CPU_ALARM_PERCENT)))

        memory = row.get("Memory Usage (%)")
        if _is_number(memory) and memory > MEMORY_ALARM_PERCENT:
            alarms.append(_alarm(time, "Fixed threshold", "Memory", memory, MEMORY_ALARM_PERCENT,
                "Memory usage {}% is above the {}% limit".format(memory, MEMORY_ALARM_PERCENT)))

        gpu = row.get("GPU Usage (%)")
        if _is_number(gpu) and gpu > GPU_ALARM_PERCENT:
            alarms.append(_alarm(time, "Fixed threshold", "GPU", gpu, GPU_ALARM_PERCENT,
                "GPU usage {}% is above the {}% limit".format(gpu, GPU_ALARM_PERCENT)))

        temperature = row.get("CPU Temperature (°C)")
        if _is_number(temperature) and temperature > TEMPERATURE_ALARM_C:
            alarms.append(_alarm(time, "Fixed threshold", "Temperature", temperature, TEMPERATURE_ALARM_C,
                "CPU temperature {}C is above the {}C limit".format(temperature, TEMPERATURE_ALARM_C)))

        battery = row.get("Battery Charge (%)")
        plugged = row.get("Power Plugged In")
        if _is_number(battery) and battery < BATTERY_LOW_PERCENT and plugged == "No":
            alarms.append(_alarm(time, "Fixed threshold", "Battery", battery, BATTERY_LOW_PERCENT,
                "Battery charge {}% is below the {}% floor while on battery power".format(
                    battery, BATTERY_LOW_PERCENT)))

    for row in read_log_sheet("disk"):
        usage = row.get("Disk Usage (%)")
        if _is_number(usage) and usage > DISK_ALARM_PERCENT:
            drive = row.get("Drive")
            alarms.append(_alarm(row.get("Time"), "Fixed threshold", "Disk " + str(drive),
                usage, DISK_ALARM_PERCENT,
                "Drive {} usage {}% is above the {}% limit".format(drive, usage, DISK_ALARM_PERCENT)))

    for row in read_log_sheet("services"):
        name = row.get("Service Name")
        status = row.get("Status")
        if name in REQUIRED_SERVICES and status != "running":
            alarms.append(_alarm(row.get("Time"), "Fixed threshold", "Service " + str(name),
                status, "running",
                "Required service {} is {} instead of running".format(name, status)))

    return alarms


import statistics

# Metrics the statistical benchmark learns a normal range for. Non-numeric readings
# such as an Unavailable temperature are skipped automatically.
BASELINE_METRICS = [
    "Overall CPU Usage (%)",
    "Memory Usage (%)",
    "GPU Usage (%)",
    "CPU Temperature (°C)",
]

# Need at least this many normal readings of a metric to trust its range.
MIN_BASELINE_SAMPLES = 3


def compute_baseline(rows, metrics=None):
    """Work out the normal range (mean and spread) for each metric from normal rows."""
    metrics = metrics if metrics is not None else BASELINE_METRICS
    baseline = {}
    for metric in metrics:
        values = [row.get(metric) for row in rows if _is_number(row.get(metric))]
        if len(values) < MIN_BASELINE_SAMPLES:
            continue
        mean = statistics.mean(values)
        spread = statistics.pstdev(values)
        baseline[metric] = {
            "mean": round(mean, 2),
            "spread": round(spread, 2),
            "low": round(mean - BASELINE_STD_MULTIPLIER * spread, 2),
            "high": round(mean + BASELINE_STD_MULTIPLIER * spread, 2),
        }
    return baseline


def check_statistical_baseline(rows=None):
    """Learn a normal range from the first part of the data and flag later outliers."""
    rows = rows if rows is not None else read_log_sheet("system_metrics")
    if len(rows) < MIN_BASELINE_SAMPLES + 1:
        return [], {}

    split = max(1, int(len(rows) * BASELINE_FRACTION))
    baseline = compute_baseline(rows[:split])
    if not baseline:
        return [], {}

    alarms = []
    for row in rows[split:]:
        time = row.get("Time")
        for metric, normal in baseline.items():
            value = row.get(metric)
            if not _is_number(value):
                continue
            if value < normal["low"] or value > normal["high"]:
                normal_range = "{} to {}".format(normal["low"], normal["high"])
                alarms.append(_alarm(time, "Statistical baseline", metric, value, normal_range,
                    "{} reading {} is outside the normal range {}".format(
                        metric, value, normal_range)))
    return alarms, baseline


import json


def save_benchmark(benchmark, path=None):
    """Write a learned normal profile to a benchmark file."""
    path = path or BENCHMARK_PATH
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(benchmark, handle, indent=2)


def load_benchmark(path=None):
    """Read a saved normal profile, or None if there is no saved file yet."""
    path = path or BENCHMARK_PATH
    if not path.exists():
        return None
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def check_saved_baseline(rows=None, path=None):
    """Compare samples against a saved benchmark, creating it the first time.

    Returns the alarms, the benchmark used, and whether it was created this call.
    """
    rows = rows if rows is not None else read_log_sheet("system_metrics")
    benchmark = load_benchmark(path)
    created = False

    if benchmark is None:
        # No saved benchmark yet. Treat the given data as normal and save it.
        benchmark = compute_baseline(rows)
        if not benchmark:
            return [], None, False
        save_benchmark(benchmark, path)
        created = True

    alarms = []
    for row in rows:
        time = row.get("Time")
        for metric, normal in benchmark.items():
            value = row.get(metric)
            if not _is_number(value):
                continue
            if value < normal["low"] or value > normal["high"]:
                normal_range = "{} to {}".format(normal["low"], normal["high"])
                alarms.append(_alarm(time, "Saved baseline", metric, value, normal_range,
                    "{} reading {} is outside the saved normal range {}".format(
                        metric, value, normal_range)))
    return alarms, benchmark, created


from collections import Counter

from openpyxl import load_workbook

ALARM_HEADERS = ["Time", "Source", "Resource", "Reading", "Limit", "Message"]


def gather_all_alarms():
    """Run all three benchmark methods and return every alarm they raise, combined."""
    fixed = check_fixed_thresholds()
    statistical, _ = check_statistical_baseline()
    saved, _, _ = check_saved_baseline()
    return fixed + statistical + saved


def write_alarms_sheet(alarms):
    """Write the combined alarms to an alarms sheet in both Excel files."""
    for path in (DASHBOARD_PATH, LOG_PATH):
        if not path.exists():
            continue
        wb = load_workbook(path)
        if "alarms" in wb.sheetnames:
            del wb["alarms"]
        ws = wb.create_sheet("alarms")
        ws.append(ALARM_HEADERS)
        for alarm in alarms:
            ws.append([alarm.get(header) for header in ALARM_HEADERS])
        wb.save(path)


def conclude_alarms():
    """Combine every method's alarms, write them to Excel, and print the verdict."""
    alarms = gather_all_alarms()
    write_alarms_sheet(alarms)

    print("=" * 60)
    print("CONCLUDING ALARM")
    print("=" * 60)
    if not alarms:
        print("ALL CLEAR. No deviations from normal were detected.")
    else:
        print("ALARM. {} deviations from normal were detected.".format(len(alarms)))
        print("")
        print("By detection method:")
        for source, count in Counter(a["Source"] for a in alarms).most_common():
            print("   {:<22} {}".format(source, count))
        print("")
        print("Most affected resources:")
        for resource, count in Counter(a["Resource"] for a in alarms).most_common(5):
            print("   {:<22} {}".format(resource, count))
    print("")
    print("All alarms written to the alarms sheet in the dashboard and the log.")
    return alarms


import matplotlib
matplotlib.use("Agg")  # draw to image files, no popup window needed
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# Columns of the trend table, in order. Column 1 is Time, the rest are the metrics.
TREND_COLUMNS = [
    "Time",
    "CPU Usage (%)",
    "Memory Usage (%)",
    "Battery Charge (%)",
    "GPU Usage (%)",
    "Data Sent (MB)",
    "Data Received (MB)",
]


def build_trend_table():
    """Read the log and build one trend row per sample with the headline metrics."""
    system = read_log_sheet("system_metrics")
    network = read_log_sheet("network")

    # Network is per interface, so add up the per-sample amounts for each timestamp.
    sent_by_time = {}
    recv_by_time = {}
    for row in network:
        time = row.get("Time")
        sent = row.get("Data Sent Since Last Sample (MB)")
        recv = row.get("Data Received Since Last Sample (MB)")
        if _is_number(sent):
            sent_by_time[time] = sent_by_time.get(time, 0) + sent
        if _is_number(recv):
            recv_by_time[time] = recv_by_time.get(time, 0) + recv

    def numeric_or_none(value):
        return value if _is_number(value) else None

    table = []
    for row in system:
        time = row.get("Time")
        table.append({
            "Time": time,
            "CPU Usage (%)": numeric_or_none(row.get("Overall CPU Usage (%)")),
            "Memory Usage (%)": numeric_or_none(row.get("Memory Usage (%)")),
            "Battery Charge (%)": numeric_or_none(row.get("Battery Charge (%)")),
            "GPU Usage (%)": numeric_or_none(row.get("GPU Usage (%)")),
            "Data Sent (MB)": round(sent_by_time.get(time, 0), 3),
            "Data Received (MB)": round(recv_by_time.get(time, 0), 3),
        })
    return table


def _save_png_charts(table):
    """Save the trend graphs as PNG image files in the charts folder."""
    if not table:
        return []
    times = [row["Time"] for row in table]
    saved = []

    def plot(filename, title, ylabel, series):
        figure, axis = plt.subplots(figsize=(10, 4))
        plotted = False
        for label, key in series:
            points = [(t, row[key]) for t, row in zip(times, table) if _is_number(row.get(key))]
            if points:
                xs, ys = zip(*points)
                axis.plot(xs, ys, marker="o", label=label)
                plotted = True
        axis.set_title(title)
        axis.set_xlabel("Time")
        axis.set_ylabel(ylabel)
        if plotted:
            axis.legend()
        figure.autofmt_xdate()
        figure.tight_layout()
        out_path = CHARTS_DIR / filename
        figure.savefig(out_path, dpi=100)
        plt.close(figure)
        saved.append(out_path)

    plot("cpu_memory.png", "CPU and Memory Usage Over Time", "Percent (%)",
         [("CPU", "CPU Usage (%)"), ("Memory", "Memory Usage (%)")])
    plot("battery.png", "Battery Charge Over Time", "Percent (%)",
         [("Battery", "Battery Charge (%)")])
    plot("gpu.png", "GPU Usage Over Time", "Percent (%)",
         [("GPU", "GPU Usage (%)")])
    plot("network.png", "Network Data Per Sample Over Time", "Megabytes (MB)",
         [("Sent", "Data Sent (MB)"), ("Received", "Data Received (MB)")])
    return saved


def _add_excel_charts(table):
    """Add native Excel line charts to the dashboard, backed by a trends data sheet."""
    if not table or not DASHBOARD_PATH.exists():
        return
    wb = load_workbook(DASHBOARD_PATH)
    for name in ("trends", "charts"):
        if name in wb.sheetnames:
            del wb[name]

    trends = wb.create_sheet("trends")
    trends.append(TREND_COLUMNS)
    for row in table:
        trends.append([row.get(column) for column in TREND_COLUMNS])

    charts = wb.create_sheet("charts")
    last_row = len(table) + 1
    categories = Reference(trends, min_col=1, min_row=2, max_row=last_row)

    def line(title, min_col, max_col, anchor):
        chart = LineChart()
        chart.title = title
        chart.height = 8
        chart.width = 16
        data = Reference(trends, min_col=min_col, max_col=max_col, min_row=1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        charts.add_chart(chart, anchor)

    # Trend columns: B=CPU, C=Memory, D=Battery, E=GPU, F=Sent, G=Received.
    line("CPU and Memory Usage (%)", 2, 3, "A1")
    line("Battery Charge (%)", 4, 4, "A18")
    line("GPU Usage (%)", 5, 5, "A35")
    line("Network Data Per Sample (MB)", 6, 7, "A52")
    wb.save(DASHBOARD_PATH)


def build_charts():
    """Build the trend graphs as PNG files and as Excel charts in the dashboard."""
    table = build_trend_table()
    saved = _save_png_charts(table)
    _add_excel_charts(table)
    print("Charts built from", len(table), "samples.")
    print("PNG images saved:", [path.name for path in saved])
    if table:
        print("Trends and charts sheets added to the dashboard.")
    return table


def run_full_monitor(interval=None, duration=None, fresh_start=True):
    """Collect samples, raise the concluding alarm, then build the charts, in order."""
    run_monitor(interval=interval, duration=duration, fresh_start=fresh_start)
    print("")
    conclude_alarms()
    print("")
    build_charts()
    print("")
    print("Done. The dashboard, log, alarms, benchmark and charts are in the outputs folder.")


if __name__ == "__main__":
    import sys
    from datetime import timedelta

    if "--demo" in sys.argv:
        run_full_monitor(interval=timedelta(seconds=10), duration=timedelta(seconds=30))
    else:
        run_full_monitor()
